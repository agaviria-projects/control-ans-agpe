from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy



def preparar_agpe_clean_excel():
    print("🧩 Aplicando validaciones Excel en AGPE_CLEAN...")

    base_dir = Path(__file__).resolve().parents[2]
    clean_dir = base_dir / "data_clean"
    ruta_excel = clean_dir / "AGPE_CLEAN.xlsx"

    if not ruta_excel.exists():
        raise FileNotFoundError("❌ No existe AGPE_CLEAN.xlsx")

    wb = load_workbook(ruta_excel)
    ws = wb.active


    # --------------------------------------------------
    # ✅ ASEGURAR COLUMNA "Tipo Visita" (SI NO EXISTE)
    # --------------------------------------------------
   # ✅ ASEGURAR COLUMNA "Tipo Visita" (SI NO EXISTE)
    encabezados = {cell.value: cell.column for cell in ws[1]}

    if "Tipo Visita" not in encabezados:
        nueva_col = ws.max_column + 1

        celda_nueva = ws.cell(row=1, column=nueva_col)
        celda_nueva.value = "Tipo Visita"

        # ✅ Copiar estilo del encabezado anterior (para que quede igual: letra negra, fondo, bordes, etc.)
        celda_ref = ws.cell(row=1, column=nueva_col - 1)

        celda_nueva.font = copy(celda_ref.font)
        celda_nueva.fill = copy(celda_ref.fill)
        celda_nueva.border = copy(celda_ref.border)
        celda_nueva.alignment = copy(celda_ref.alignment)
        celda_nueva.number_format = celda_ref.number_format
        celda_nueva.protection = copy(celda_ref.protection)

        # ancho base opcional
        ws.column_dimensions[celda_nueva.column_letter].width = 12

        # refrescar encabezados para que el resto del script la detecte
        encabezados = {cell.value: cell.column for cell in ws[1]}

    # --------------------------------------------------
    # 1️⃣ Buscar columna DETALLE_VISITA
    # --------------------------------------------------
    encabezados = {cell.value: cell.column for cell in ws[1]}

    if "Detalle Visita" not in encabezados:
        raise ValueError("❌ No se encontró la columna Detalle Visita")

    col_detalle = encabezados["Detalle Visita"]
    letra_col = ws.cell(row=1, column=col_detalle).column_letter

    # --------------------------------------------------
    # 2️⃣ Crear hoja LISTAS (Opción B)
    # --------------------------------------------------
    if "LISTAS" not in wb.sheetnames:
        ws_listas = wb.create_sheet("LISTAS")
    else:
        ws_listas = wb["LISTAS"]

    ws_listas["A1"] = "Detalle Visita"

    ws_listas["B1"] = "Tipo Visita"

    opciones_tipo_visita = ["C07", "C08", "C09"]

    for i, opcion in enumerate(opciones_tipo_visita, start=2):
        ws_listas[f"B{i}"] = opcion


    opciones_detalle = [
        "1ER VISITA",
        "2DA VISITA Y DOCUMENTOS",
        "3ER VISITA Y DOCUMENTOS",
        "4TA VISITA Y DOCUMENTOS",
        "5TA VISITA Y DOCUMENTOS",
        "DOCUMENTOS",
        "DIRECTA",
        "SEMIDIRECTA",
        "INDIRECTA"
    ]

    for i, opcion in enumerate(opciones_detalle, start=2):
        ws_listas[f"A{i}"] = opcion

    ws_listas.sheet_state = "hidden"

    # --------------------------------------------------
    # 3️⃣ Definir validación desde hoja LISTAS
    # --------------------------------------------------
    validacion = DataValidation(
        type="list",
        formula1="=LISTAS!$A$2:$A$10",
        allow_blank=True,
        showDropDown=False
    )

    validacion.promptTitle = "Detalle Visita"
    validacion.prompt = "Seleccione un valor de la lista"
    validacion.errorTitle = "Valor no permitido"
    validacion.error = "Debe seleccionar un valor válido de la lista."

    # ws.add_data_validation(validacion)

    # --------------------------------------------------
    # 3️⃣ Aplicar validación a filas existentes + margen
    # --------------------------------------------------
    ultima_fila = ws.max_row
    fila_inicio = 2
    fila_fin = max(ultima_fila + 500, 1000)

    rango = f"{letra_col}{fila_inicio}:{letra_col}{fila_fin}"
    # validacion.add(rango)

    # --------------------------------------------------
    # ✅ VALIDACIÓN: TIPO VISITA (C07, C08, C09)
    # --------------------------------------------------
    encabezados = {cell.value: cell.column for cell in ws[1]}
    if "Tipo Visita" in encabezados:
        col_tv = encabezados["Tipo Visita"]
        letra_tv = ws.cell(row=1, column=col_tv).column_letter

        dv_tv = DataValidation(
            type="list",
            formula1="=LISTAS!$B$2:$B$4",
            allow_blank=True,
            showDropDown=False
        )
        dv_tv.promptTitle = "Tipo Visita"
        dv_tv.prompt = "Seleccione C07, C08 o C09"
        dv_tv.errorTitle = "Valor no permitido"
        dv_tv.error = "Debe seleccionar un valor válido (C07, C08, C09)."

        ws.add_data_validation(dv_tv)
        dv_tv.add(f"{letra_tv}2:{letra_tv}1048576")


    # --------------------------------------------------
    # B3️⃣ Congelar encabezados + ajuste de ancho
    # --------------------------------------------------
    ws.freeze_panes = "A2"

    ancho_max = 45
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, ancho_max)
    
    # --------------------------------------------------
    # B4️⃣ Proteger columnas técnicas
    # --------------------------------------------------
    columnas_editables = [
        "Detalle Visita",
        "Tipo Medidor",
        "Tipo Visita"

    ]

    encabezados = {cell.value: cell.column for cell in ws[1]}

    # Bloquear todo
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = cell.protection.copy(locked=True)

    # Desbloquear solo columnas permitidas
    for col_nombre in columnas_editables:
        if col_nombre in encabezados:
            col_idx = encabezados[col_nombre]
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=col_idx).protection = ws.cell(
                    row=fila, column=col_idx
                ).protection.copy(locked=False)

    # Permitir seleccionar solo celdas desbloqueadas
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True

    # Activar protección
    # ws.protection.enable()

    # --------------------------------------------------
    # B5️⃣ Convertir AGPE_CLEAN en tabla estructurada
    # --------------------------------------------------
    if ws.tables:
        ws.tables.clear()

    max_fila = ws.max_row
    max_col = ws.max_column
    letra_col_final = ws.cell(row=1, column=max_col).column_letter

    rango_tabla = f"A1:{letra_col_final}{max_fila}"

    tabla = Table(displayName="tbl_AGPE_CLEAN", ref=rango_tabla)

    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    # --------------------------------------------------
    # B7️⃣ Aplicar VALIDACIÓN DE DATOS (DESPUÉS de la tabla)
    # --------------------------------------------------

    # from openpyxl.worksheet.datavalidation import DataValidation

    validacion = DataValidation(
        type="list",
        formula1="=LISTAS!$A$2:$A$10",
        allow_blank=True,
        showDropDown=False
    )

    validacion.promptTitle = "Detalle Visita"
    validacion.prompt = "Seleccione un valor de la lista"
    validacion.errorTitle = "Valor no permitido"
    validacion.error = "Debe seleccionar un valor válido."

    ws.add_data_validation(validacion)

    # Aplicar a toda la columna Detalle Visita (desde fila 2)
    ultima_fila = ws.max_row
    rango = f"{letra_col}2:{letra_col}1048576"
    validacion.add(rango)

    # --------------------------------------------------
    # B6️⃣ Proteger columnas técnicas (DESPUÉS de la tabla)
    # --------------------------------------------------

    columnas_editables = [
        "Detalle Visita",
        "Tipo Medidor"
    ]

    encabezados = {cell.value: cell.column for cell in ws[1]}

    # Bloquear todo
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = cell.protection.copy(locked=True)

    # Desbloquear solo columnas editables
    for col_nombre in columnas_editables:
        if col_nombre in encabezados:
            col_idx = encabezados[col_nombre]
            for fila in range(2, ws.max_row + 1):
                ws.cell(row=fila, column=col_idx).protection = ws.cell(
                    row=fila, column=col_idx
                ).protection.copy(locked=False)

    # Configurar y activar protección correctamente
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    # ws.protection.enable()

    # --------------------------------------------------
    # 🔓 Dejar hoja SIN protección para el usuario final
    # --------------------------------------------------
    ws.protection.disable()

    # --------------------------------------------------
    # 4️⃣ Guardar
    # --------------------------------------------------
    wb.save(ruta_excel)
    wb.close()

    print("✅ Validación Detalle Visita aplicada correctamente")