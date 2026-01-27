print("🔥 EJECUTANDO append_agpe_ans.py 🔥")

from pathlib import Path
import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date
from pandas.tseries.offsets import CustomBusinessDay


# ============================================================
# PASO 0) UTILIDADES SEGURAS (NO CRÍTICAS)
# ============================================================

def _norm_headers(cols):
    return (
        pd.Series(cols)
        .astype(str)
        .str.strip()
        .str.upper()
        .str.replace(" ", "_", regex=False)
        .str.replace("Á", "A")
        .str.replace("É", "E")
        .str.replace("Í", "I")
        .str.replace("Ó", "O")
        .str.replace("Ú", "U")
        .tolist()
    )

def _safe_str(s):
    if s is None or (isinstance(s, float) and np.isnan(s)):
        return ""
    return str(s).strip()

def _dias_ans_por_detalle(detalle):
    d = _safe_str(detalle).upper().strip()

    # 12 días hábiles: "1ER/1RA/2DA/3ER/3RA/4TA/5TA VISITA" (aunque NO diga DOCUMENTOS)
    if "VISITA" in d and any(x in d for x in ("1ER", "1RA", "2DA", "3ER", "3RA", "4TA", "5TA")):
        return 12

    # 12 días hábiles: "X VISITA Y DOCUMENTOS" o similares
    if "VISITA" in d and ("DOC" in d or "DOCUMENT" in d):
        return 12

    # 5 días hábiles
    if d == "DOCUMENTOS" or d.startswith("DOCUMENT"):
        return 5

    # 9 días hábiles
    if d in ("DIRECTA", "SEMIDIRECTA", "INDIRECTA"):
        return 9

    return None


def _dias_ans_por_tipo_visita(tipo_visita):
    tv = str(tipo_visita).upper().strip()
    
    if tv in ("C08", "C09"):
        return 9 
    return None

def _estado_por_dias_restantes(dias):
    if dias is None:
        return ""
    if dias < 0:
        return "VENCIDO"
    if dias == 0:
        return "ALERTA 0 DIAS"
    if dias <= 2:
        return "ALERTA"
    return "A TIEMPO"


def _find_col_letter_by_header(ws, header_name_upper):
    header_name_upper = str(header_name_upper).strip().upper()
    for cell in ws[1]:
        if str(cell.value).strip().upper() == header_name_upper:
            return cell.column_letter
    return None

def _find_col_index_by_header(ws, header_name_upper):
    header_name_upper = str(header_name_upper).strip().upper()
    for cell in ws[1]:
        if str(cell.value).strip().upper() == header_name_upper:
            return cell.col_idx
    return None

def _update_or_create_table(ws, table_name="tbl_AGPE_ANS"):
    # Actualiza el rango de la tabla sin destruir formatos
    max_row = ws.max_row
    max_col = ws.max_column
    last_col_letter = ws.cell(row=1, column=max_col).column_letter
    ref = f"A1:{last_col_letter}{max_row}"

    if table_name in ws.tables:
        ws.tables[table_name].ref = ref
    else:
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(t)

def _ensure_obs_validation(ws, max_row, formula_range="=validador!$A$1:$A$9"):
    """
    Aplica validación de lista a OBSERVACION para todo el rango actual.
    """
    col_obs = _find_col_letter_by_header(ws, "OBSERVACION")
    if not col_obs:
        return

    dv = DataValidation(
        type="list",
        formula1=formula_range,
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.add(f"{col_obs}2:{col_obs}{max_row}")

def _limpiar_excel_dejar_encabezados_xlsx(ruta_xlsx):
    """
    Deja un .xlsx en blanco conservando encabezados.
    """
    df_vacio = pd.read_excel(ruta_xlsx, nrows=0)
    df_vacio.to_excel(ruta_xlsx, index=False)

def _limpiar_excel_dejar_encabezados_xlsm(ruta_xlsm):
    """
    Borra filas (2..fin) en la hoja activa de un .xlsm conservando macros.
    """
    wb = load_workbook(ruta_xlsm, keep_vba=True)
    ws = wb.active
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    wb.save(ruta_xlsm)
    wb.close()


# ============================================================
# PASO 1) DEFINIR COLUMNAS CONTROLADAS POR APPEND (sin tocar W)
# ============================================================

COLUMNAS_MIN_CONTROLADAS = [
    "PEDIDO",
    "DIRECCION",
    "MUNICIPIO",
    "CLIENTE",
    "SUBZONA",
    "PROMOTOR",
    "CELULAR",
    "FECHA_ATENCION",
    "HORA_VISITA",
    "REVISOR",
    "TIPO_VISITA",
    "OBSERVACION",
    "FECHA_CIERRE_FENIX",
    "POTENCIA_AC_KW",
    "DETALLE_VISITA",
    "COORDENADAX",
    "COORDENADAY",
    "TIPO_MEDIDOR",
    "URBANO_RURAL",
    "FECHA_CAMBIO_ESTADO",
    "FECHA_LIMITE_ANS",
    "DIAS_RESTANTES",
    "ESTADO_ANS",  # NO se toca, pero debe existir en la plantilla
]


# ============================================================
# PASO 2) APPEND PRINCIPAL
# ============================================================

def append_agpe_ans():
    print("➡️ Iniciando APPEND seguro a AGPE_ANS (sin calendario)")

    base_dir = Path(__file__).resolve().parents[2]
    ruta_clean = base_dir / "data_clean" / "AGPE_CLEAN.xlsx"
    ruta_visitas = base_dir / "data_clean" / "PLANTILLA_PRIMER VISITAS.xlsm"
    ruta_ans = base_dir / "data_clean" / "AGPE_ANS.xlsm"
    ruta_bdpcp = base_dir / "data_clean" / "BDPCP.xlsx"

    if not ruta_ans.exists():
        raise FileNotFoundError("❌ No existe AGPE_ANS.xlsm")
    if not ruta_clean.exists():
        raise FileNotFoundError("❌ No existe AGPE_CLEAN.xlsx")
    if not ruta_visitas.exists():
        raise FileNotFoundError("❌ No existe PLANTILLA_PRIMER VISITAS.xlsm")

    # ============================================================
    # PASO 3) ABRIR AGPE_ANS + ELIMINAR FILA 2 SI ESTÁ VACÍA
    # ============================================================

    wb_ans = load_workbook(ruta_ans, keep_vba=True, data_only=False)
    ws_ans = wb_ans.active

    def _eliminar_fila2_si_vacia(ws):
        """
        Elimina la fila 2 SOLO si está completamente vacía (todas las celdas vacías/None).
        Esto evita que quede esa “fila azul” vacía debajo del encabezado.
        """
        if ws.max_row < 2:
            return

        hay_algo = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=2, column=c).value
            if _safe_str(v) != "":
                hay_algo = True
                break

        if not hay_algo:
            ws.delete_rows(2, 1)

    # ✅ Quitar la fila 2 “vacía azul” si existe
    _eliminar_fila2_si_vacia(ws_ans)

    col_pedido_idx = _find_col_index_by_header(ws_ans, "PEDIDO")
    if not col_pedido_idx:
        wb_ans.close()
        raise ValueError("❌ AGPE_ANS no tiene columna PEDIDO en encabezados.")

    pedidos_historicos = set()
    for r in range(2, ws_ans.max_row + 1):
        v = ws_ans.cell(row=r, column=col_pedido_idx).value
        v = _safe_str(v).upper()
        if v != "":
            pedidos_historicos.add(v)

    print(f"📌 PEDIDOS Históricos en AGPE_ANS: {len(pedidos_historicos)}")

    # Encabezados reales de AGPE_ANS (para respetar orden)
    headers_ans = []
    for c in range(1, ws_ans.max_column + 1):
        v = ws_ans.cell(row=1, column=c).value
        headers_ans.append(str(v).strip().upper() if v is not None else "")
    headers_ans = [h for h in headers_ans if h != ""]

    if "PEDIDO" not in headers_ans:
        wb_ans.close()
        raise ValueError("❌ Encabezados inválidos en AGPE_ANS: falta PEDIDO.")

    # ============================================================
    # PASO 4) LEER AGPE_CLEAN (NUEVOS) Y NORMALIZAR
    # ============================================================

    df_clean = pd.read_excel(ruta_clean, dtype=str)
    df_clean.columns = _norm_headers(df_clean.columns)

    mapeo_clean = {
        "NOMBRE_CLIENTE": "CLIENTE",
        "FECHA_INICIO_ANS": "FECHA_CAMBIO_ESTADO",
        "TIPO_DIRECCIÓN": "URBANO_RURAL",
        "URBANO/RURAL": "URBANO_RURAL",
        "TIPO_DIRECCION": "URBANO_RURAL",
        "TIPO_DE_MEDIDOR": "TIPO_MEDIDOR",
        "MEDIDOR": "TIPO_MEDIDOR",
        "TIPO_MEDIDOR_": "TIPO_MEDIDOR",
    }
    for k, v in list(mapeo_clean.items()):
        if k in df_clean.columns:
            df_clean = df_clean.rename(columns={k: v})

    if "TIPO_MEDIDOR" not in df_clean.columns:
        df_clean["TIPO_MEDIDOR"] = ""
    if "URBANO_RURAL" not in df_clean.columns:
        df_clean["URBANO_RURAL"] = ""
    if "DETALLE_VISITA" not in df_clean.columns:
        df_clean["DETALLE_VISITA"] = ""
    if "ACTIVIDAD" not in df_clean.columns:
        df_clean["ACTIVIDAD"] = ""

    req_clean = ["PEDIDO", "DIRECCION", "MUNICIPIO", "CLIENTE", "SUBZONA", "COORDENADAX", "COORDENADAY", "FECHA_CAMBIO_ESTADO"]
    faltantes_clean = [c for c in req_clean if c not in df_clean.columns]
    if faltantes_clean:
        wb_ans.close()
        raise ValueError(f"❌ Columnas faltantes en AGPE_CLEAN: {faltantes_clean}")

    df_clean["PEDIDO"] = df_clean["PEDIDO"].astype(str).str.strip().str.upper()
    df_clean["DIRECCION"] = df_clean["DIRECCION"].astype(str).str.strip().str.lstrip("'")
    df_clean["CLIENTE"] = df_clean["CLIENTE"].astype(str).str.strip().str.upper()
    df_clean["SUBZONA"] = (
        df_clean["SUBZONA"].astype(str).str.strip().str.upper()
        .replace({
            "METROPOLITANA SUR": "METROPOLITANA",
            "METROPOLITANA-SUR": "METROPOLITANA",
            "METROPOLITANA  SUR": "METROPOLITANA",
        })
    )
    df_clean["DETALLE_VISITA"] = df_clean["DETALLE_VISITA"].astype(str).str.strip().str.upper()
    df_clean["URBANO_RURAL"] = df_clean["URBANO_RURAL"].astype(str).str.strip().str.upper()

    # df_clean["TIPO_VISITA"] = df_clean["ACTIVIDAD"].apply(
    #     lambda x: "C09" if str(x).strip().upper() == "ACVIS" else "C07"
    # )

    df_clean_datos = df_clean[df_clean["PEDIDO"].notna() & (df_clean["PEDIDO"].str.strip() != "")]

    # ============================================================
    # PASO 5) LEER PRIMERAS VISITAS (NUEVOS) Y NORMALIZAR
    # ============================================================

    df_vis = pd.read_excel(ruta_visitas, sheet_name=0, dtype=str)
    df_vis.columns = _norm_headers(df_vis.columns)

    if "PEDIDO" not in df_vis.columns:
        wb_ans.close()
        raise ValueError("❌ PLANTILLA_PRIMER VISITAS no tiene columna PEDIDO.")

    df_vis["PEDIDO"] = df_vis["PEDIDO"].astype(str).str.strip().str.upper()
    df_vis_datos = df_vis[df_vis["PEDIDO"].notna() & (df_vis["PEDIDO"].str.strip() != "")]

    if "SUBZONA" not in df_vis.columns:
        df_vis["SUBZONA"] = ""
    if "SUBZONA_ID" in df_vis.columns:
        map_subzona = {
            "ORI": "ORIENTE",
            "MET": "METROPOLITANA",
            "OCC": "OCCIDENTE",
            "SUR": "SUROESTE",
            "ND": "NORDESTE",
        }
        df_vis["SUBZONA"] = (
            df_vis["SUBZONA_ID"].astype(str).str.strip().str.upper().map(map_subzona).fillna("")
        )

    # ============================================================
    # PASO 6) SALIDA TEMPRANA CONTROLADA
    # ============================================================

    if df_clean_datos.empty and df_vis_datos.empty:
        wb_ans.close()
        raise RuntimeError("AGPE_CLEAN y PLANTILLA_PRIMER VISITAS no contienen datos para procesar.")

    # ============================================================
    # PASO 7) CONSTRUIR FILAS NUEVAS (SIN TOCAR W, SIN CALENDARIO)
    # ============================================================

    def _blank_row_dict():
        d = {h: "" for h in headers_ans}
        for h in COLUMNAS_MIN_CONTROLADAS:
            if h not in d:
                d[h] = ""
        return d

    nuevas_filas = []

    for _, row in df_clean_datos.iterrows():
        pedido = _safe_str(row.get("PEDIDO", "")).upper()
        if pedido == "":
            continue

        d = _blank_row_dict()
        d["PEDIDO"] = pedido
        d["DIRECCION"] = _safe_str(row.get("DIRECCION", ""))
        d["MUNICIPIO"] = _safe_str(row.get("MUNICIPIO", ""))
        d["CLIENTE"] = _safe_str(row.get("CLIENTE", "")).upper()
        d["SUBZONA"] = _safe_str(row.get("SUBZONA", "")).upper()
        d["TIPO_VISITA"] = _safe_str(row.get("TIPO_VISITA", ""))
        d["DETALLE_VISITA"] = _safe_str(row.get("DETALLE_VISITA", "")).upper()
        if "1RA VISITA" in d["DETALLE_VISITA"]:
            d["TIPO_VISITA"] = "C07"
        d["COORDENADAX"] = _safe_str(row.get("COORDENADAX", ""))
        d["COORDENADAY"] = _safe_str(row.get("COORDENADAY", ""))
        d["TIPO_MEDIDOR"] = _safe_str(row.get("TIPO_MEDIDOR", ""))
        d["URBANO_RURAL"] = _safe_str(row.get("URBANO_RURAL", "")).upper()
        d["FECHA_CAMBIO_ESTADO"] = _safe_str(row.get("FECHA_CAMBIO_ESTADO", ""))

        nuevas_filas.append(d)
        

    for _, row in df_vis_datos.iterrows():
        pedido = _safe_str(row.get("PEDIDO", "")).upper()
        if pedido == "" or pedido in pedidos_historicos:
            continue

        d = _blank_row_dict()
        d["PEDIDO"] = pedido
        d["DIRECCION"] = _safe_str(row.get("DIRECCION", ""))
        d["MUNICIPIO"] = _safe_str(row.get("MUNICIPIO", ""))
        d["CLIENTE"] = _safe_str(row.get("CLIENTE", "")).upper()
        d["SUBZONA"] = _safe_str(row.get("SUBZONA", "")).upper()

        d["PROMOTOR"] = _safe_str(row.get("PROMOTOR", ""))
        d["CELULAR"] = _safe_str(row.get("CELULAR", ""))
        d["FECHA_ATENCION"] = ""
        d["HORA_VISITA"] = ""
        d["POTENCIA_AC_KW"] = _safe_str(row.get("POTENCIA_AC_KW", ""))

        d["DETALLE_VISITA"] = _safe_str(row.get("DETALLE_VISITA", "")).upper()
        d["COORDENADAX"] = _safe_str(row.get("COORDENADAX", ""))
        d["COORDENADAY"] = _safe_str(row.get("COORDENADAY", ""))
        d["URBANO_RURAL"] = _safe_str(row.get("URBANO_RURAL", "")).upper()
        d["TIPO_VISITA"] = _safe_str(row.get("TIPO_VISITA", "")).upper()
        d["OBSERVACION"] = _safe_str(row.get("OBSERVACION", "")).upper()
        d["FECHA_CAMBIO_ESTADO"] = _safe_str(row.get("FECHA_CAMBIO_ESTADO", ""))

        nuevas_filas.append(d)

    if not nuevas_filas:
        wb_ans.close()
        raise RuntimeError("No se encontraron pedidos nuevos para agregar (todo ya existe en AGPE_ANS).")

    print(f"✅ Filas NUEVAS a agregar: {len(nuevas_filas)}")

    # ============================================================
    # PASO 8) CARGAR BDPCP Y CONSTRUIR MAPA
    # ============================================================

    bdpcp_map = {}
    if ruta_bdpcp.exists():
        df_bdpcp = pd.read_excel(ruta_bdpcp, dtype=str)
        df_bdpcp.columns = _norm_headers(df_bdpcp.columns)

        # 🔧 ALIAS SEGURO PARA POTENCIA (CORRECCIÓN MÍNIMA)
        if "POTENCIA_AC_[KW]" in df_bdpcp.columns:
            df_bdpcp = df_bdpcp.rename(columns={"POTENCIA_AC_[KW]": "POTENCIA_AC_KW"})

        req_bd = ["PEDIDO", "PROMOTOR", "CELULAR", "POTENCIA_AC_KW"]
        if all(c in df_bdpcp.columns for c in req_bd):
            df_bdpcp["PEDIDO"] = df_bdpcp["PEDIDO"].astype(str).str.strip().str.upper()
            for _, r in df_bdpcp.iterrows():
                p = _safe_str(r.get("PEDIDO", "")).upper()
                if p:
                    bdpcp_map[p] = {
                        "PROMOTOR": _safe_str(r.get("PROMOTOR", "")),
                        "CELULAR": _safe_str(r.get("CELULAR", "")),
                        "POTENCIA_AC_KW": _safe_str(r.get("POTENCIA_AC_KW", "")),
                    }

    # ===== DEBUG TEMPORAL (NO AFECTA LÓGICA) =====
    print("DEBUG BDPCP keys:", list(bdpcp_map.keys())[:5])
    print("DEBUG pedido ejemplo:", list(pedidos_historicos)[:5])

    # ============================================================
    # PASO 8.2) CRUZAR BDPCP TAMBIÉN EN FILAS EXISTENTES (solo llena vacíos)
    # ============================================================

    if bdpcp_map:
        idx_prom = _find_col_index_by_header(ws_ans, "PROMOTOR")
        idx_cel = _find_col_index_by_header(ws_ans, "CELULAR")
        idx_pot = _find_col_index_by_header(ws_ans, "POTENCIA_AC_KW")

        # Si alguna columna no existe, no rompemos: solo saltamos ese relleno.
        for r in range(2, ws_ans.max_row + 1):
            pedido = _safe_str(ws_ans.cell(row=r, column=col_pedido_idx).value).upper()
            if not pedido or pedido not in bdpcp_map:
                continue

            if idx_prom:
                v = _safe_str(ws_ans.cell(row=r, column=idx_prom).value)
                if v == "":
                    ws_ans.cell(row=r, column=idx_prom).value = bdpcp_map[pedido].get("PROMOTOR", "")

            if idx_cel:
                v = _safe_str(ws_ans.cell(row=r, column=idx_cel).value)
                if v == "":
                    ws_ans.cell(row=r, column=idx_cel).value = bdpcp_map[pedido].get("CELULAR", "")

            if idx_pot:
                v = _safe_str(ws_ans.cell(row=r, column=idx_pot).value)
                if v == "":
                    ws_ans.cell(row=r, column=idx_pot).value = bdpcp_map[pedido].get("POTENCIA_AC_KW", "")
    # ============================================================
    # PASO 8.1) CRUZAR BDPCP EN NUEVAS FILAS (solo llena vacíos)
    # ============================================================

    for d in nuevas_filas:
        p = _safe_str(d.get("PEDIDO", "")).upper()
        if p in bdpcp_map:
            if _safe_str(d.get("PROMOTOR", "")) == "":
                d["PROMOTOR"] = bdpcp_map[p].get("PROMOTOR", "")
            if _safe_str(d.get("CELULAR", "")) == "":
                d["CELULAR"] = bdpcp_map[p].get("CELULAR", "")
            if _safe_str(d.get("POTENCIA_AC_KW", "")) == "":
                d["POTENCIA_AC_KW"] = bdpcp_map[p].get("POTENCIA_AC_KW", "")
    # ============================================================
    # PASO 9) APPEND REAL AL FINAL (SIN BORRAR HISTÓRICO)
    # ============================================================

    start_row = ws_ans.max_row + 1

    for d in nuevas_filas:
        row_values = [d.get(h, "") for h in headers_ans]
        ws_ans.append(row_values)

    end_row = ws_ans.max_row
    print(f"📌 AGPE_ANS: filas agregadas desde {start_row} hasta {end_row}")

    # ============================================================
    # PASO 9.1) CALCULAR ANS SOLO PARA FILAS NUEVAS
    # ============================================================

    idx_fecha_cambio = _find_col_index_by_header(ws_ans, "FECHA_CAMBIO_ESTADO")
    idx_detalle = _find_col_index_by_header(ws_ans, "DETALLE_VISITA")
    idx_fecha_limite = _find_col_index_by_header(ws_ans, "FECHA_LIMITE_ANS")
    idx_dias_rest = _find_col_index_by_header(ws_ans, "DIAS_RESTANTES")
    idx_estado = _find_col_index_by_header(ws_ans, "ESTADO_ANS")
    idx_tipo_visita = _find_col_index_by_header(ws_ans, "TIPO_VISITA")

    if not all([idx_fecha_cambio, idx_detalle, idx_fecha_limite, idx_dias_rest, idx_estado]):
        wb_ans.close()
        raise ValueError("❌ Faltan columnas clave para cálculo ANS (U, V, W).")
    
    hoy = date.today()
    bd = CustomBusinessDay()

    for r in range(start_row, end_row + 1):

        detalle = ws_ans.cell(r, idx_detalle).value

        # 🔧 Regla: si DETALLE_VISITA es 1RA VISITA => TIPO_VISITA = C07
        detalle_norm = _safe_str(detalle).upper()
        if "1RA VISITA" in detalle_norm:
            tipo_visita = "C07"
            if idx_tipo_visita:
                ws_ans.cell(r, idx_tipo_visita).value = "C07"

        # 🔒 NO TOCAR SI YA EXISTE VALOR (protección absoluta)
        if _safe_str(ws_ans.cell(r, idx_fecha_limite).value) != "":
            continue

        fecha_cambio_raw = ws_ans.cell(r, idx_fecha_cambio).value
        detalle = ws_ans.cell(r, idx_detalle).value

        if not fecha_cambio_raw:
            continue

        try:
            fecha_cambio = pd.to_datetime(fecha_cambio_raw)
        except Exception:
            continue

        # fallback solo si DETALLE no aplica
        dias_ans = _dias_ans_por_detalle(detalle)
        if dias_ans is None:
            dias_ans = _dias_ans_por_tipo_visita(tipo_visita)
        if dias_ans is None:
            continue

        # ✅ NUEVA LÓGICA (SOLO AQUÍ)
        tipo_visita = ws_ans.cell(
            r,
            _find_col_index_by_header(ws_ans, "TIPO_VISITA")
        ).value

        # ✅ PRIORIDAD CORRECTA: DETALLE manda
        dias_ans = _dias_ans_por_detalle(detalle)

        fecha_limite = fecha_cambio + CustomBusinessDay(n=dias_ans)
        dias_restantes = (fecha_limite.date() - hoy).days
        estado = _estado_por_dias_restantes(dias_restantes)

        ws_ans.cell(r, idx_fecha_limite).value = fecha_limite.to_pydatetime()
        ws_ans.cell(r, idx_dias_rest).value = dias_restantes
        ws_ans.cell(r, idx_estado).value = estado



    # ============================================================
    # PASO 10) MANTENER TABLA + VALIDACIÓN (SIN TOCAR FORMATO CONDICIONAL)
    # ============================================================

    _update_or_create_table(ws_ans, table_name="tbl_AGPE_ANS")
    _ensure_obs_validation(ws_ans, max_row=ws_ans.max_row, formula_range="=validador!$A$1:$A$9")

    wb_ans.save(ruta_ans)
    wb_ans.close()

    # ============================================================
    # PASO 11) LIMPIAR FUENTES (SOLO SI TODO SALIÓ BIEN)
    # ============================================================

    _limpiar_excel_dejar_encabezados_xlsx(ruta_clean)
    _limpiar_excel_dejar_encabezados_xlsm(ruta_visitas)

    print("✅ APPEND finalizado. AGPE_ANS actualizado, BDPCP cruzado y fuentes limpiadas.")


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    append_agpe_ans()