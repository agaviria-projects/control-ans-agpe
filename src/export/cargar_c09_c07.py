from pathlib import Path
from openpyxl import load_workbook


def limpiar_plantilla_c09_c07():
    """
    Limpia la PLANTILLA C09_C07.xlsx eliminando filas de datos.
    NO borra encabezados.
    NO afecta validaciones ni formatos.
    """

    base_dir = Path(__file__).resolve().parents[2]
    ruta_plantilla = base_dir / "entrada_diaria" / "PLANTILLA C09_C07.xlsx"

    if not ruta_plantilla.exists():
        raise FileNotFoundError("No existe PLANTILLA C09_C07.xlsx en entrada_diaria")

    print("🧹 Abriendo plantilla C09_C07 (modo seguro)...")

    wb = load_workbook(ruta_plantilla)
    ws = wb.active

    fila_inicio_datos = 2
    total_filas = ws.max_row

    if total_filas < fila_inicio_datos:
        print("ℹ️ La plantilla ya está vacía.")
    else:
        filas_a_eliminar = total_filas - fila_inicio_datos + 1
        print(f"🧽 Eliminando {filas_a_eliminar} filas de datos...")
        ws.delete_rows(fila_inicio_datos, filas_a_eliminar)

    wb.save(ruta_plantilla)
    wb.close()

    print("✅ Plantilla C09_C07 limpia correctamente (sin afectar validaciones).")